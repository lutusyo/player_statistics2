from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.drawing.image import Image as ExcelImage
from players_app.models import Player
from teams_app.models import Team
from django.conf import settings
import os

def export_team_players_to_excel_view(request, team_id):
    team = Team.objects.get(id=team_id)

    wb = Workbook()
    ws = wb.active
    ws.title = f"{team.name} Players"

    headers = [
        "NAME", "JNAME", "AGE GROUP", "BIRTHDATE",
        "POSITION", "HEIGHT (CM)", "WEIGHT (KG)", "BMI",
        "FOOT PREFERENCE", "JERSEY NUMBER", "FORMER CLUB",
        "PLAYER PHONE", "PARENT PHONE", "PHOTO"
    ]

    # Styles
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")

    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    players = Player.objects.filter(team=team).select_related("age_group")

    for player in players:
        latest_measurement = player.current_measurement
        height = latest_measurement.height if latest_measurement else ""
        weight = latest_measurement.weight if latest_measurement else ""
        bmi = round(float(weight)/(float(height)/100)**2, 1) if height and weight else ""

        full_name = f"{player.name} {player.second_name} {player.surname}".upper()

        row = [
            full_name,
            (player.jina_maarufu or "").upper(),
            player.age_group.name.upper() if player.age_group else "",
            player.birthdate.strftime("%Y-%m-%d") if player.birthdate else "",
            (player.position or "").upper(),
            height,
            weight,
            bmi,
            (player.foot_preference or "").upper(),
            player.jersey_number or "",
            (player.former_club or "").upper(),
            player.player_phone or "",
            player.parent_phone or "",
            ""  # Placeholder for photo
        ]

        ws.append(row)

        # Add player photo
        if player.photo and os.path.exists(os.path.join(settings.MEDIA_ROOT, player.photo.name)):
            img_path = os.path.join(settings.MEDIA_ROOT, player.photo.name)
            img = ExcelImage(img_path)
            img.width = 40
            img.height = 40
            ws.add_image(img, f"N{ws.max_row}")  # Last column

    # Align all cells
    for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
        for cell in row_cells:
            cell.alignment = center_align

    # Adjust column widths
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 5

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"{team.name.replace(' ', '_')}_PLAYERS.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response
