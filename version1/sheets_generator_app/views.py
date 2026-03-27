from django.http import HttpResponse
from django.shortcuts import render
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side

def lineup_template(request):
    return render(request, "sheets_generator_app/lineup_template.html")
    


def download_lineup_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Lineup Sheet"

    # Styles
    bold = Font(bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Header
    ws.merge_cells('A1:C1')
    ws['A1'] = "AZAM FOOTBALL CLUB ACADEMY"
    ws['A1'].font = bold
    ws['A1'].alignment = Alignment(horizontal="center")

    ws.merge_cells('A2:C2')
    ws['A2'] = "MATCH LINE-UP SHEET"
    ws['A2'].font = bold
    ws['A2'].alignment = Alignment(horizontal="center")

    # Match info
    ws['A4'] = "Competition:"
    ws['B4'] = ""
    ws['A5'] = "Age Group:"
    ws['B5'] = ""
    ws['A6'] = "Date:"
    ws['B6'] = ""
    ws['A7'] = "Kickoff Time:"
    ws['B7'] = ""
    ws['A8'] = "Opponent:"
    ws['B8'] = ""
    ws['A9'] = "Pitch No:"
    ws['B9'] = ""

    # Starting XI header
    ws['A11'] = "STARTING XI"
    ws['A11'].font = bold

    ws.append(["No.", "Player Name", "Position"])
    for cell in ws[12]:
        cell.font = bold
        cell.border = thin_border

    # 11 empty rows
    for _ in range(11):
        ws.append(["________", "__________________________________", "____________"])
        for cell in ws[ws.max_row]:
            cell.border = thin_border

    # Substitutes section
    ws.append([])
    ws.append(["SUBSTITUTES"])
    ws[ws.max_row][0].font = bold

    ws.append(["No.", "Player Name"])
    for cell in ws[ws.max_row]:
        cell.font = bold
        cell.border = thin_border

    for _ in range(10):
        ws.append(["________", "___________________________________"])
        for cell in ws[ws.max_row]:
            cell.border = thin_border

    # Staff
    ws.append([])
    ws.append(["Coach Name:", "______________________________________"])

    # Output Excel
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="lineup_sheet.xlsx"'

    wb.save(response)
    return response
