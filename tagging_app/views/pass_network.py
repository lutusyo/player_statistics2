# tagging_app/views/pass_network.py
from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse, HttpResponse, FileResponse
from django.views.decorators.csrf import csrf_exempt

from collections import defaultdict
import csv
import io
import json

from django.db.models import Count
from players_app.models import Player
from matches_app.models import Match
from lineup_app.models import MatchLineup
from tagging_app.models import PassEvent
from teams_app.models import Team

# ReportLab (PDF)
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle, SimpleDocTemplate, Paragraph, Image, Spacer
from reportlab.lib.styles import getSampleStyleSheet

# Excel
from openpyxl import Workbook

# Matplotlib (server-side)
import matplotlib
matplotlib.use('Agg')  # non-GUI backend for servers
import matplotlib.pyplot as plt



# ---------- SHARED DATA FUNCTION ----------
def get_pass_network_data(match_id):
    """
    Returns:
        players: QuerySet of Player (ordered by name)
        player_names: dict {player_id: player_name}
        matrix: dict[from_id][to_id] = count
        total_passes: dict[from_id] = total passes made
        ball_lost: dict[from_id] = passes that changed team away from passer's team
    """

    match = get_object_or_404(Match, id=match_id)
    player_ids_from = PassEvent.objects.filter(match=match).values_list('from_player_id', flat=True)
    player_ids_to = PassEvent.objects.filter(match=match).values_list('to_player_id', flat=True)
    all_player_ids = set(player_ids_from).union(set(player_ids_to))

    players = Player.objects.filter(id__in=all_player_ids).order_by('name')
    player_names = {p.id: p.name for p in players}

    # Build counts
    matrix = defaultdict(lambda: defaultdict(int))
    passes = (
        PassEvent.objects.filter(match=match)
        .values('from_player_id', 'to_player_id', 'from_team_id', 'to_team_id')
        .annotate(count=Count('id'))
    )

    total_passes = defaultdict(int)
    ball_lost = defaultdict(int)

    for p in passes:
        from_id = p['from_player_id']
        to_id = p['to_player_id']
        cnt = p['count']

        if to_id:
            matrix[from_id][to_id] = cnt
        total_passes[from_id] += cnt
        if p['from_team_id'] != p['to_team_id']:
            ball_lost[from_id] += cnt

    return players, player_names, matrix, total_passes, ball_lost



def pass_network_enter_data(request, match_id):
    players, player_names, matrix, total_passes, ball_lost = get_pass_network_data(match_id)

    # Get team information (optional based on your template)
    match = get_object_or_404(Match, id=match_id)
    our_team = match.home_team  # Or determine based on logic
    opponent_team = match.away_team

    # Pass all required data to the template
    context = {
        'match': match,
        'our_players': [{'player': p} for p in players],  # Wrap like this to match template structure
        'our_team': our_team,
        'opponent_team': opponent_team,
        'matrix': matrix,
        'total_passes': total_passes,
        'ball_lost': ball_lost,
    }

    return render(request, 'tagging_app/pass_network_enter_data.html', context)



# ---------- SAVE ----------
@csrf_exempt
def save_pass_network(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        try:
            match = Match.objects.get(id=data['match_id'])
            from_player = Player.objects.get(id=data['from_player_id'])
            to_player = Player.objects.get(id=data.get('to_player_id')) if data.get('to_player_id') else None
            from_team = Team.objects.get(id=data['from_team_id'])
            to_team = Team.objects.get(id=data['to_team_id'])

            PassEvent.objects.create(
                match=match,
                from_player=from_player,
                to_player=to_player,
                from_team=from_team,
                to_team=to_team,
                minute=data['minute'],
                second=data['second'],
                x_start=data.get('x_start'),
                y_start=data.get('y_start'),
                x_end=data.get('x_end'),
                y_end=data.get('y_end'),
                is_successful=data['is_successful'],
                is_possession_regained=data.get('is_possession_regained', False),
            )

            count = PassEvent.objects.filter(match=match, from_player=from_player).count()
            return JsonResponse({'status': 'success', 'count': count})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})

    return JsonResponse({'status': 'invalid request'})


# ---------- DASHBOARD (HTML) ----------
def pass_network_dashboard(request, match_id):
    match = get_object_or_404(Match, id=match_id)
    players, player_names, matrix, total_passes, ball_lost = get_pass_network_data(match_id)

    # Top 5 combinations
    combos = []
    for f_id, tos in matrix.items():
        for t_id, c in tos.items():
            combos.append((
                player_names.get(f_id, f"Unknown({f_id})"),
                player_names.get(t_id, f"Unknown({t_id})"),
                c
            ))
    top_combinations = sorted(combos, key=lambda x: x[2], reverse=True)[:5]

    # Bar chart data
    chart_data = {
        'names': [player_names[p.id] for p in players],
        'values': [total_passes[p.id] for p in players]
    }

    context = {
        'match': match,
        'players': players,
        'matrix': matrix,
        'player_names': player_names,
        'top_combinations': top_combinations,
        'total_passes': total_passes,
        'ball_lost': ball_lost,
        'chart_data': chart_data,
    }
    return render(request, 'tagging_app/pass_network_dashboard.html', context)



# ---------- CSV ----------
def export_pass_network_csv(request, match_id):
    match = get_object_or_404(Match, id=match_id)
    players, player_names, matrix, total_passes, ball_lost = get_pass_network_data(match_id)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="pass_network_{match_id}.csv"'
    writer = csv.writer(response)

    writer.writerow([f"Passing Matrix - {match}"])
    header = ["From\\To"] + [player_names[p.id][:10] for p in players]
    writer.writerow(header)
    for row_player in players:
        row = [player_names[row_player.id][:10]]
        for col_player in players:
            row.append("—" if row_player.id == col_player.id else str(matrix[row_player.id].get(col_player.id, 0)))
        writer.writerow(row)

    writer.writerow([])
    writer.writerow(["Player", "Total Passes", "Ball Lost"])
    for p in players:
        writer.writerow([player_names[p.id], total_passes[p.id], ball_lost[p.id]])

    return response


# ---------- EXCEL ----------
def export_pass_network_excel(request, match_id):
    match = get_object_or_404(Match, id=match_id)
    players, player_names, matrix, total_passes, ball_lost = get_pass_network_data(match_id)

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Pass Matrix"
    ws1.cell(row=1, column=1, value="From\\To")
    for col_index, p in enumerate(players, start=2):
        ws1.cell(row=1, column=col_index, value=player_names[p.id][:10])

    for row_index, from_player in enumerate(players, start=2):
        ws1.cell(row=row_index, column=1, value=player_names[from_player.id][:10])
        for col_index, to_player in enumerate(players, start=2):
            ws1.cell(row=row_index, column=col_index,
                     value="—" if from_player.id == to_player.id else matrix[from_player.id].get(to_player.id, 0))

    ws2 = wb.create_sheet(title="Totals")
    ws2.append(["Player", "Total Passes", "Ball Lost"])
    for p in players:
        ws2.append([player_names[p.id], total_passes[p.id], ball_lost[p.id]])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="pass_network_{match_id}.xlsx"'
    wb.save(response)
    return response


# ---------- PDF ----------
def export_pass_network_pdf(request, match_id):
    match = get_object_or_404(Match, id=match_id)
    players, player_names, matrix, total_passes, ball_lost = get_pass_network_data(match_id)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()

    # Title
    elements.append(Paragraph(f"Passing Network - {match}", styles['Title']))
    elements.append(Spacer(1, 8))

    # Top 5 Combinations
    combos = []
    for f_id, tos in matrix.items():
        for t_id, c in tos.items():
            combos.append((player_names.get(f_id, str(f_id)), player_names.get(t_id, str(t_id)), c))
    top_combinations = sorted(combos, key=lambda x: x[2], reverse=True)[:5]

    elements.append(Paragraph("Top 5 Pass Combinations", styles['Heading2']))
    combo_data = [["From", "To", "Count"]]
    for f, t, c in top_combinations:
        combo_data.append([f, t, str(c)])
    combo_table = Table(combo_data)
    combo_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.lightblue),
        ('GRID', (0,0), (-1,-1), 0.5, colors.black),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTSIZE', (0,0), (-1,-1), 9),
    ]))
    elements.append(combo_table)
    elements.append(Spacer(1,12))

    # Pass Matrix
    data = [["From\\To"] + [player_names[p.id][:10] for p in players]]
    for row_player in players:
        row = [player_names[row_player.id][:10]]
        for col_player in players:
            row.append("—" if row_player.id == col_player.id else str(matrix[row_player.id].get(col_player.id, 0)))
        data.append(row)

    t = Table(data)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.blue),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.black),
        ('FONTSIZE', (0,0), (-1,-1), 9),
    ]))
    elements.append(Spacer(1,12))
    elements.append(Paragraph("Pass Matrix", styles['Heading2']))
    elements.append(t)
    elements.append(Spacer(1,12))

    # Totals Table
    totals_data = [["Player", "Total Passes", "Ball Lost"]]
    for p in players:
        totals_data.append([player_names[p.id], str(total_passes[p.id]), str(ball_lost[p.id])])
    totals_table = Table(totals_data)
    totals_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.darkblue),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.black),
        ('FONTSIZE', (0,0), (-1,-1), 9),
    ]))
    elements.append(Paragraph("Player Passing Totals", styles['Heading2']))
    elements.append(totals_table)
    elements.append(Spacer(1,16))

    # Bar Chart
    if players:
        fig, ax = plt.subplots(figsize=(6, 4))
        names = [player_names[p.id] for p in players]
        values = [total_passes[p.id] for p in players]
        ax.bar(names, values)
        ax.set_ylabel("Total Passes")
        ax.set_title("Passes per Player")
        plt.xticks(rotation=45, ha='right')
        plt.tight_layout()

        img_buf = io.BytesIO()
        plt.savefig(img_buf, format='png')
        plt.close(fig)
        img_buf.seek(0)

        elements.append(Paragraph("Pass Distribution Chart", styles['Heading2']))
        elements.append(Image(img_buf, width=400, height=250))

    doc.build(elements)
    buffer.seek(0)
    return FileResponse(buffer, as_attachment=True, filename='pass_network.pdf')
