from django.contrib.auth.decorators import login_required
from django.db.models import Count, Q
from django.http import HttpResponse
from django.shortcuts import render
from django.utils import timezone

from apps.medical_data.forms.medical_filter_form import MedicalFilterForm
from apps.medical_data.models.medical_visit import MedicalVisit
from apps.medical_data.models.medical_follow_up import MedicalFollowUp
from apps.medical_data.models.medical_recovery_plan import (
    MedicalRecoveryPlan,
    RecoveryPlanStatus,
)

# CHANGE THIS IMPORT PATH IF YOUR TRAINING MODELS ARE IN ANOTHER APP
from version1.reports_app.models.previous_models import (
    TrainingMinutes,
    PlayerTrainingMinutes,
    TrainingAbsence,
)

from apps.core.choices import AvailabilityStatus

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter






from io import BytesIO

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    PageBreak,
)







# ============================================================
# FILTER MEDICAL VISITS
# ============================================================

def get_filtered_medical_visits(request):

    queryset = (
        MedicalVisit.objects
        .select_related("team", "player", "created_by")
        .order_by("-date")
    )

    form = MedicalFilterForm(request.GET or None)

    if form.is_valid():

        filters = {
            "date__gte": form.cleaned_data.get("start_date"),
            "date__lte": form.cleaned_data.get("end_date"),
            "team": form.cleaned_data.get("team"),
            "player": form.cleaned_data.get("player"),
            "visit_type": form.cleaned_data.get("visit_type"),
            "main_complaint": form.cleaned_data.get("main_complaint"),
            "availability_status": form.cleaned_data.get(
                "availability_status"
            ),
        }

        filters = {
            key: value
            for key, value in filters.items()
            if value not in (None, "")
        }

        queryset = queryset.filter(**filters)

    return queryset, form


# ============================================================
# REPORT DATA
# ============================================================

def get_medical_report_data(request):

    medical_visits, filter_form = get_filtered_medical_visits(request)

    today = timezone.localdate()

    # --------------------------------------------------------
    # FILTER VALUES
    # --------------------------------------------------------

    start_date = None
    end_date = None
    selected_team = None
    selected_player = None

    if filter_form.is_valid():

        start_date = filter_form.cleaned_data.get("start_date")
        end_date = filter_form.cleaned_data.get("end_date")
        selected_team = filter_form.cleaned_data.get("team")
        selected_player = filter_form.cleaned_data.get("player")


    # ========================================================
    # MEDICAL SUMMARY
    # ========================================================

    total_records = medical_visits.count()

    new_injuries = medical_visits.filter(
        visit_type="new_injury"
    ).count()

    regular_checkups = medical_visits.filter(
        visit_type="regular_checkup"
    ).count()

    available_players = (
        medical_visits
        .filter(availability_status=AvailabilityStatus.AVAILABLE)
        .values("player")
        .distinct()
        .count()
    )

    restricted_players = (
        medical_visits
        .filter(availability_status=AvailabilityStatus.RESTRICTED)
        .values("player")
        .distinct()
        .count()
    )

    unavailable_players = (
        medical_visits
        .filter(availability_status=AvailabilityStatus.NOT_AVAILABLE)
        .values("player")
        .distinct()
        .count()
    )


    # ========================================================
    # COMPLAINT ANALYSIS
    # ========================================================

    complaint_data = (
        medical_visits
        .values("main_complaint")
        .annotate(total=Count("id"))
        .order_by("-total")
    )


    # ========================================================
    # TEAM ANALYSIS
    # ========================================================

    team_data = (
        medical_visits
        .values("team__name")
        .annotate(total=Count("id"))
        .order_by("-total")
    )

    team_labels = [
        item["team__name"]
        for item in team_data
    ]

    team_values = [
        item["total"]
        for item in team_data
    ]


    # ========================================================
    # TRAINING ABSENCES
    # ========================================================

    absence_queryset = TrainingAbsence.objects.select_related(
        "player",
        "training_session",
        "training_session__team",
    )

    if start_date:
        absence_queryset = absence_queryset.filter(
            training_session__date__gte=start_date
        )

    if end_date:
        absence_queryset = absence_queryset.filter(
            training_session__date__lte=end_date
        )

    if selected_team:
        absence_queryset = absence_queryset.filter(
            training_session__team=selected_team
        )

    if selected_player:
        absence_queryset = absence_queryset.filter(
            player=selected_player
        )


    # ========================================================
    # TRAINING ABSENCE SUMMARY
    # ========================================================

    total_training_absences = absence_queryset.count()

    injured_absences = absence_queryset.filter(
        reason="INJURED"
    ).count()

    sick_absences = absence_queryset.filter(
        reason="SICK"
    ).count()

    personal_absences = absence_queryset.filter(
        reason="PERSONAL"
    ).count()

    unexcused_absences = absence_queryset.filter(
        reason="UNEXCUSED"
    ).count()


    # ========================================================
    # TRAINING MISSED BY PLAYER
    # ========================================================

    training_absence_by_player = (
        absence_queryset
        .values(
            "player",
            "training_session__team__name",
        )
        .annotate(
            total_missed=Count("id"),

            injured=Count(
                "id",
                filter=Q(reason="INJURED"),
            ),

            sick=Count(
                "id",
                filter=Q(reason="SICK"),
            ),

            personal=Count(
                "id",
                filter=Q(reason="PERSONAL"),
            ),

            unexcused=Count(
                "id",
                filter=Q(reason="UNEXCUSED"),
            ),
        )
        .order_by("-total_missed")
    )


    # ========================================================
    # RECENT / FOLLOW-UP DATA
    # ========================================================

    followup_queryset = MedicalFollowUp.objects.select_related(
        "visit",
        "visit__player",
        "visit__team",
    )

    if start_date:
        followup_queryset = followup_queryset.filter(
            visit__date__gte=start_date
        )

    if end_date:
        followup_queryset = followup_queryset.filter(
            visit__date__lte=end_date
        )

    if selected_team:
        followup_queryset = followup_queryset.filter(
            visit__team=selected_team
        )

    if selected_player:
        followup_queryset = followup_queryset.filter(
            visit__player=selected_player
        )


    overdue_followups = followup_queryset.filter(
        status=False,
        review_date__lt=today,
    ).order_by("review_date")


    upcoming_followups = followup_queryset.filter(
        status=False,
        review_date__gte=today,
    ).order_by("review_date")


    # ========================================================
    # RECOVERY PLANS
    # ========================================================

    recovery_queryset = MedicalRecoveryPlan.objects.select_related(
        "visit",
        "visit__player",
        "visit__team",
    )

    if start_date:
        recovery_queryset = recovery_queryset.filter(
            start_date__gte=start_date
        )

    if end_date:
        recovery_queryset = recovery_queryset.filter(
            start_date__lte=end_date
        )

    if selected_team:
        recovery_queryset = recovery_queryset.filter(
            visit__team=selected_team
        )

    if selected_player:
        recovery_queryset = recovery_queryset.filter(
            visit__player=selected_player
        )


    active_recovery_plans = recovery_queryset.filter(
        status=RecoveryPlanStatus.ACTIVE
    ).count()

    completed_recovery_plans = recovery_queryset.filter(
        status=RecoveryPlanStatus.COMPLETED
    ).count()

    cancelled_recovery_plans = recovery_queryset.filter(
        status=RecoveryPlanStatus.CANCELLED
    ).count()


    # ========================================================
    # EXPECTED RETURNS
    # ========================================================

    upcoming_returns = (
        medical_visits
        .filter(
            expected_return_date__isnull=False,
            expected_return_date__gte=today,
        )
        .order_by("expected_return_date")
    )


    # ========================================================
    # RECENT MEDICAL VISITS
    # ========================================================

    recent_visits = medical_visits[:10]


    # ========================================================
    # FINAL CONTEXT
    # ========================================================

    context = {

        "page_title": "Medical Report",

        # Filters
        "filter_form": filter_form,

        # Medical summary
        "medical_visits": medical_visits,
        "total_records": total_records,
        "new_injuries": new_injuries,
        "regular_checkups": regular_checkups,
        "available_players": available_players,
        "restricted_players": restricted_players,
        "unavailable_players": unavailable_players,

        # Analysis
        "complaint_data": complaint_data,
        "team_labels": team_labels,
        "team_values": team_values,
        "team_data": team_data,

        # Training impact
        "total_training_absences": total_training_absences,
        "injured_absences": injured_absences,
        "sick_absences": sick_absences,
        "personal_absences": personal_absences,
        "unexcused_absences": unexcused_absences,
        "training_absence_by_player": training_absence_by_player,

        # Recovery
        "active_recovery_plans": active_recovery_plans,
        "completed_recovery_plans": completed_recovery_plans,
        "cancelled_recovery_plans": cancelled_recovery_plans,

        # Follow-ups
        "overdue_followups": overdue_followups,
        "upcoming_followups": upcoming_followups,

        # Returns
        "upcoming_returns": upcoming_returns,

        # Recent visits
        "recent_visits": recent_visits,
    }

    return context


# ============================================================
# MEDICAL REPORT PAGE
# ============================================================

@login_required
def medical_report(request):

    context = get_medical_report_data(request)

    return render(
        request,
        "medical_data/medical_report.html",
        context,
    )




@login_required
def medical_report_excel(request):
    context = get_medical_report_data(request)

    wb = Workbook()

    # ==========================================================
    # STYLES
    # ==========================================================

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78"
    )

    section_fill = PatternFill(
        fill_type="solid",
        fgColor="D9EAF7"
    )

    white_font = Font(
        color="FFFFFF",
        bold=True
    )

    bold_font = Font(
        bold=True
    )

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def style_header(ws, row):
        for cell in ws[row]:
            cell.fill = header_fill
            cell.font = white_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )
            cell.border = thin_border

    def auto_width(ws):
        for column_cells in ws.columns:
            max_length = 0
            column = get_column_letter(column_cells[0].column)

            for cell in column_cells:
                try:
                    value_length = len(str(cell.value or ""))
                    max_length = max(max_length, value_length)
                except Exception:
                    pass

            ws.column_dimensions[column].width = min(
                max(max_length + 2, 12),
                40
            )

    def add_title(ws, title):
        ws["A1"] = title
        ws["A1"].font = Font(
            bold=True,
            size=16
        )
        ws["A1"].alignment = Alignment(
            horizontal="center"
        )

        ws.merge_cells(
            start_row=1,
            start_column=1,
            end_row=1,
            end_column=6
        )

    # ==========================================================
    # SHEET 1 — MEDICAL SUMMARY
    # ==========================================================

    ws = wb.active
    ws.title = "Medical Summary"

    add_title(ws, "Medical Report Summary")

    ws["A3"] = "Report Information"
    ws["A3"].font = bold_font
    ws["A3"].fill = section_fill

    ws["A4"] = "Start Date"
    ws["B4"] = context["filter_form"].cleaned_data.get("start_date") \
        if context["filter_form"].is_valid() else ""

    ws["A5"] = "End Date"
    ws["B5"] = context["filter_form"].cleaned_data.get("end_date") \
        if context["filter_form"].is_valid() else ""

    ws["A6"] = "Team"
    ws["B6"] = context["filter_form"].cleaned_data.get("team") \
        if context["filter_form"].is_valid() else ""

    ws["A7"] = "Player"
    ws["B7"] = context["filter_form"].cleaned_data.get("player") \
        if context["filter_form"].is_valid() else ""

    ws["A9"] = "Medical Statistics"
    ws["A9"].font = bold_font
    ws["A9"].fill = section_fill

    summary_data = [
        ("Total Medical Records", context["total_records"]),
        ("New Injuries", context["new_injuries"]),
        ("Regular Checkups", context["regular_checkups"]),
        ("Available Players", context["available_players"]),
        ("Restricted Players", context["restricted_players"]),
        ("Unavailable Players", context["unavailable_players"]),
        ("Active Recovery Plans", context["active_recovery_plans"]),
        ("Completed Recovery Plans", context["completed_recovery_plans"]),
        ("Cancelled Recovery Plans", context["cancelled_recovery_plans"]),
        ("Total Training Absences", context["total_training_absences"]),
        ("Injury-related Absences", context["injured_absences"]),
        ("Sick Absences", context["sick_absences"]),
        ("Personal Absences", context["personal_absences"]),
        ("Unexcused Absences", context["unexcused_absences"]),
    ]

    row = 10

    for label, value in summary_data:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)

        ws.cell(row=row, column=1).font = bold_font

        ws.cell(row=row, column=1).border = thin_border
        ws.cell(row=row, column=2).border = thin_border

        row += 1

    auto_width(ws)

    # ==========================================================
    # SHEET 2 — INJURY ANALYSIS
    # ==========================================================

    ws = wb.create_sheet("Injury Analysis")

    add_title(ws, "Medical Complaint Analysis")

    headers = [
        "Main Complaint",
        "Total Records",
    ]

    ws.append([])
    ws.append(headers)

    style_header(ws, 3)

    for item in context["complaint_data"]:
        ws.append([
            item["main_complaint"],
            item["total"],
        ])

    for row in ws.iter_rows(min_row=4):
        for cell in row:
            cell.border = thin_border

    auto_width(ws)

    # ==========================================================
    # SHEET 3 — TRAINING IMPACT
    # ==========================================================

    ws = wb.create_sheet("Training Impact")

    add_title(ws, "Training Impact")

    ws.append([])
    ws.append([
        "Absence Reason",
        "Training Sessions Missed",
    ])

    style_header(ws, 3)

    training_impact = [
        ("Injured", context["injured_absences"]),
        ("Sick", context["sick_absences"]),
        ("Personal", context["personal_absences"]),
        ("Unexcused", context["unexcused_absences"]),
        ("Total", context["total_training_absences"]),
    ]

    for reason, total in training_impact:
        ws.append([
            reason,
            total,
        ])

    for row in ws.iter_rows(min_row=4):
        for cell in row:
            cell.border = thin_border

    auto_width(ws)

    # ==========================================================
    # SHEET 4 — PLAYER TRAINING IMPACT
    # ==========================================================

    ws = wb.create_sheet("Player Training Impact")

    add_title(ws, "Training Sessions Missed by Player")

    ws.append([])
    ws.append([
        "Player",
        "Team",
        "Total Missed",
        "Injured",
        "Sick",
        "Personal",
        "Unexcused",
    ])

    style_header(ws, 3)

    for item in context["training_absence_by_player"]:

        ws.append([
            str(item["player"]),
            item["training_session__team__name"],
            item["total_missed"],
            item["injured"],
            item["sick"],
            item["personal"],
            item["unexcused"],
        ])

    for row in ws.iter_rows(min_row=4):
        for cell in row:
            cell.border = thin_border

    auto_width(ws)

    # ==========================================================
    # SHEET 5 — RECOVERY & FOLLOW-UPS
    # ==========================================================

    ws = wb.create_sheet("Recovery & Follow-ups")

    add_title(ws, "Recovery Plans and Follow-ups")

    # Recovery plans

    ws["A3"] = "Recovery Plans"
    ws["A3"].font = bold_font
    ws["A3"].fill = section_fill

    recovery_headers = [
        "Player",
        "Team",
        "Start Date",
        "Expected End Date",
        "Actual Recovery Date",
        "Status",
    ]

    for col, header in enumerate(recovery_headers, start=1):
        ws.cell(row=4, column=col, value=header)

    style_header(ws, 4)

    recovery_row = 5

    recovery_plans = MedicalRecoveryPlan.objects.select_related(
        "visit",
        "visit__player",
        "visit__team",
    )

    filter_form = context["filter_form"]

    if filter_form.is_valid():

        start_date = filter_form.cleaned_data.get("start_date")
        end_date = filter_form.cleaned_data.get("end_date")
        selected_team = filter_form.cleaned_data.get("team")
        selected_player = filter_form.cleaned_data.get("player")

        if start_date:
            recovery_plans = recovery_plans.filter(
                start_date__gte=start_date
            )

        if end_date:
            recovery_plans = recovery_plans.filter(
                start_date__lte=end_date
            )

        if selected_team:
            recovery_plans = recovery_plans.filter(
                visit__team=selected_team
            )

        if selected_player:
            recovery_plans = recovery_plans.filter(
                visit__player=selected_player
            )

    for plan in recovery_plans.order_by("-start_date"):

        ws.cell(recovery_row, 1, str(plan.visit.player))
        ws.cell(recovery_row, 2, str(plan.visit.team))
        ws.cell(recovery_row, 3, plan.start_date)
        ws.cell(recovery_row, 4, plan.expected_end_date)
        ws.cell(recovery_row, 5, plan.actual_recovery_date)
        ws.cell(recovery_row, 6, plan.get_status_display())

        recovery_row += 1

    # ==========================================================
    # FOLLOW-UPS
    # ==========================================================

    followup_start = recovery_row + 2

    ws.cell(
        followup_start,
        1,
        "Follow-ups"
    )

    ws.cell(
        followup_start,
        1
    ).font = bold_font

    ws.cell(
        followup_start,
        1
    ).fill = section_fill

    followup_header_row = followup_start + 1

    followup_headers = [
        "Player",
        "Team",
        "Review Date",
        "Status",
    ]

    for col, header in enumerate(followup_headers, start=1):
        ws.cell(
            followup_header_row,
            col,
            header
        )

    style_header(ws, followup_header_row)

    followup_row = followup_header_row + 1


    # Get all pending follow-ups
    all_followups = MedicalFollowUp.objects.select_related(
        "visit",
        "visit__player",
        "visit__team",
    ).filter(
        status=False
    )


    # Apply the same report filters
    if filter_form.is_valid():

        start_date = filter_form.cleaned_data.get("start_date")
        end_date = filter_form.cleaned_data.get("end_date")
        selected_team = filter_form.cleaned_data.get("team")
        selected_player = filter_form.cleaned_data.get("player")

        if start_date:
            all_followups = all_followups.filter(
                visit__date__gte=start_date
            )

        if end_date:
            all_followups = all_followups.filter(
                visit__date__lte=end_date
            )

        if selected_team:
            all_followups = all_followups.filter(
                visit__team=selected_team
            )

        if selected_player:
            all_followups = all_followups.filter(
                visit__player=selected_player
            )


    all_followups = all_followups.order_by("review_date")


    for followup in all_followups:

        ws.cell(
            followup_row,
            1,
            str(followup.visit.player)
        )

        ws.cell(
            followup_row,
            2,
            str(followup.visit.team)
        )

        ws.cell(
            followup_row,
            3,
            followup.review_date
        )

        ws.cell(
            followup_row,
            4,
            "Completed" if followup.status else "Pending"
        )

        followup_row += 1

    # ==========================================================
    # SHEET 6 — DETAILED MEDICAL RECORDS
    # ==========================================================

    ws = wb.create_sheet("Medical Records")

    add_title(ws, "Detailed Medical Records")

    headers = [
        "Date",
        "Team",
        "Player",
        "Visit Type",
        "Main Complaint",
        "Body Side",
        "Injury Status",
        "Mechanism of Injury",
        "Training Status",
        "Availability",
        "Working Diagnosis",
        "Therapy",
        "Recommendations",
        "Next Review Date",
        "Expected Return Date",
    ]

    ws.append([])
    ws.append(headers)

    style_header(ws, 3)

    for visit in context["medical_visits"]:

        ws.append([
            visit.date,
            str(visit.team),
            str(visit.player),
            visit.get_visit_type_display(),
            visit.get_main_complaint_display(),
            visit.get_body_side_display(),
            visit.get_injury_status_display(),
            visit.get_mechanism_of_injury_display(),
            visit.get_training_session_status_display(),
            visit.get_availability_status_display(),
            visit.working_diagnosis,
            visit.therapy,
            visit.recommendations,
            visit.next_review_date,
            visit.expected_return_date,
        ])

    for row in ws.iter_rows(min_row=4):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True
            )

    auto_width(ws)

    # ==========================================================
    # FREEZE PANES
    # ==========================================================

    for sheet in wb.worksheets:
        sheet.freeze_panes = "A4"

    # ==========================================================
    # RESPONSE
    # ==========================================================

    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )

    response["Content-Disposition"] = (
        'attachment; filename="medical_report.xlsx"'
    )

    wb.save(response)

    return response









@login_required
def medical_report_pdf(request):

    context = get_medical_report_data(request)

    # ==========================================================
    # PDF SETUP
    # ==========================================================

    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=12 * mm,
        leftMargin=12 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
        title="Medical Report",
        author="Azam FC Academy",
    )

    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "MedicalTitle",
        parent=styles["Title"],
        fontSize=18,
        leading=22,
        alignment=TA_CENTER,
        spaceAfter=8,
    )

    subtitle_style = ParagraphStyle(
        "MedicalSubtitle",
        parent=styles["Normal"],
        fontSize=9,
        leading=12,
        alignment=TA_CENTER,
        textColor=colors.grey,
        spaceAfter=12,
    )

    section_style = ParagraphStyle(
        "SectionTitle",
        parent=styles["Heading2"],
        fontSize=12,
        leading=15,
        spaceBefore=8,
        spaceAfter=6,
    )

    normal_style = ParagraphStyle(
        "NormalMedical",
        parent=styles["Normal"],
        fontSize=8,
        leading=10,
    )

    small_style = ParagraphStyle(
        "SmallMedical",
        parent=styles["Normal"],
        fontSize=7,
        leading=9,
    )

    story = []

    # ==========================================================
    # HELPER FUNCTIONS
    # ==========================================================

    def make_table(data, col_widths=None, header=True):

        table = Table(
            data,
            colWidths=col_widths,
            repeatRows=1 if header else 0,
        )

        table_style = [
            ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 5),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]

        if header:
            table_style.extend([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ])

        table.setStyle(TableStyle(table_style))

        return table

    def display_value(value):
        if value in (None, ""):
            return "-"
        return str(value)

    # ==========================================================
    # TITLE
    # ==========================================================

    story.append(
        Paragraph(
            "AZAM FC ACADEMY",
            title_style,
        )
    )

    story.append(
        Paragraph(
            "MEDICAL REPORT",
            title_style,
        )
    )

    story.append(
        Paragraph(
            f"Generated on {timezone.localdate().strftime('%d %B %Y')}",
            subtitle_style,
        )
    )

    # ==========================================================
    # FILTER INFORMATION
    # ==========================================================

    story.append(
        Paragraph(
            "Report Filters",
            section_style,
        )
    )

    filter_form = context["filter_form"]

    start_date = None
    end_date = None
    selected_team = None
    selected_player = None

    if filter_form.is_valid():
        start_date = filter_form.cleaned_data.get("start_date")
        end_date = filter_form.cleaned_data.get("end_date")
        selected_team = filter_form.cleaned_data.get("team")
        selected_player = filter_form.cleaned_data.get("player")

    filter_data = [
        ["Filter", "Selected Value"],
        [
            "Start Date",
            display_value(start_date),
        ],
        [
            "End Date",
            display_value(end_date),
        ],
        [
            "Team",
            display_value(selected_team),
        ],
        [
            "Player",
            display_value(selected_player),
        ],
    ]

    story.append(
        make_table(
            filter_data,
            col_widths=[
                45 * mm,
                130 * mm,
            ],
        )
    )

    story.append(Spacer(1, 8))

    # ==========================================================
    # MEDICAL SUMMARY
    # ==========================================================

    story.append(
        Paragraph(
            "Medical Summary",
            section_style,
        )
    )

    summary_data = [
        ["Indicator", "Total"],
        [
            "Total Medical Records",
            context["total_records"],
        ],
        [
            "New Injuries",
            context["new_injuries"],
        ],
        [
            "Regular Checkups",
            context["regular_checkups"],
        ],
        [
            "Available Players",
            context["available_players"],
        ],
        [
            "Restricted Players",
            context["restricted_players"],
        ],
        [
            "Unavailable Players",
            context["unavailable_players"],
        ],
        [
            "Active Recovery Plans",
            context["active_recovery_plans"],
        ],
        [
            "Completed Recovery Plans",
            context["completed_recovery_plans"],
        ],
        [
            "Cancelled Recovery Plans",
            context["cancelled_recovery_plans"],
        ],
    ]

    story.append(
        make_table(
            summary_data,
            col_widths=[
                130 * mm,
                45 * mm,
            ],
        )
    )

    # ==========================================================
    # TRAINING IMPACT
    # ==========================================================

    story.append(
        Paragraph(
            "Training Impact",
            section_style,
        )
    )

    training_data = [
        ["Absence Reason", "Sessions Missed"],
        [
            "Injured",
            context["injured_absences"],
        ],
        [
            "Sick",
            context["sick_absences"],
        ],
        [
            "Personal",
            context["personal_absences"],
        ],
        [
            "Unexcused",
            context["unexcused_absences"],
        ],
        [
            "TOTAL",
            context["total_training_absences"],
        ],
    ]

    story.append(
        make_table(
            training_data,
            col_widths=[
                130 * mm,
                45 * mm,
            ],
        )
    )

    # ==========================================================
    # MAIN COMPLAINTS
    # ==========================================================

    story.append(
        Paragraph(
            "Medical Complaint Analysis",
            section_style,
        )
    )

    complaint_data = [
        ["Main Complaint", "Total Records"]
    ]

    for item in context["complaint_data"]:

        complaint_data.append([
            item["main_complaint"],
            item["total"],
        ])

    if len(complaint_data) == 1:
        complaint_data.append([
            "No records",
            0,
        ])

    story.append(
        make_table(
            complaint_data,
            col_widths=[
                130 * mm,
                45 * mm,
            ],
        )
    )

    # ==========================================================
    # TEAM ANALYSIS
    # ==========================================================

    story.append(
        Paragraph(
            "Team Analysis",
            section_style,
        )
    )

    team_data = [
        ["Team", "Medical Records"]
    ]

    for item in context["team_data"]:

        team_data.append([
            display_value(item["team__name"]),
            item["total"],
        ])

    if len(team_data) == 1:
        team_data.append([
            "No records",
            0,
        ])

    story.append(
        make_table(
            team_data,
            col_widths=[
                130 * mm,
                45 * mm,
            ],
        )
    )

    # ==========================================================
    # PAGE BREAK
    # ==========================================================

    story.append(PageBreak())

    # ==========================================================
    # TRAINING ABSENCES BY PLAYER
    # ==========================================================

    story.append(
        Paragraph(
            "Training Sessions Missed by Player",
            section_style,
        )
    )

    player_absence_data = [
        [
            "Player",
            "Team",
            "Total",
            "Injured",
            "Sick",
            "Personal",
            "Unexcused",
        ]
    ]

    for item in context["training_absence_by_player"]:

        player_absence_data.append([
            display_value(item["player"]),
            display_value(item["training_session__team__name"]),
            item["total_missed"],
            item["injured"],
            item["sick"],
            item["personal"],
            item["unexcused"],
        ])

    if len(player_absence_data) == 1:
        player_absence_data.append([
            "No records",
            "-",
            0,
            0,
            0,
            0,
            0,
        ])

    story.append(
        make_table(
            player_absence_data,
            col_widths=[
                38 * mm,
                30 * mm,
                18 * mm,
                18 * mm,
                18 * mm,
                20 * mm,
                23 * mm,
            ],
        )
    )

    # ==========================================================
    # RECOVERY PLANS
    # ==========================================================

    story.append(
        Paragraph(
            "Recovery Plans",
            section_style,
        )
    )

    recovery_queryset = MedicalRecoveryPlan.objects.select_related(
        "visit",
        "visit__player",
        "visit__team",
    )

    if filter_form.is_valid():

        if start_date:
            recovery_queryset = recovery_queryset.filter(
                start_date__gte=start_date
            )

        if end_date:
            recovery_queryset = recovery_queryset.filter(
                start_date__lte=end_date
            )

        if selected_team:
            recovery_queryset = recovery_queryset.filter(
                visit__team=selected_team
            )

        if selected_player:
            recovery_queryset = recovery_queryset.filter(
                visit__player=selected_player
            )

    recovery_data = [
        [
            "Player",
            "Team",
            "Start",
            "Expected End",
            "Actual Recovery",
            "Status",
        ]
    ]

    for plan in recovery_queryset.order_by("-start_date"):

        recovery_data.append([
            display_value(plan.visit.player),
            display_value(plan.visit.team),
            display_value(plan.start_date),
            display_value(plan.expected_end_date),
            display_value(plan.actual_recovery_date),
            display_value(plan.get_status_display()),
        ])

    if len(recovery_data) == 1:
        recovery_data.append([
            "No recovery plans",
            "-",
            "-",
            "-",
            "-",
            "-",
        ])

    story.append(
        make_table(
            recovery_data,
            col_widths=[
                35 * mm,
                30 * mm,
                25 * mm,
                30 * mm,
                30 * mm,
                25 * mm,
            ],
        )
    )

    # ==========================================================
    # FOLLOW-UPS
    # ==========================================================

    story.append(
        Paragraph(
            "Pending Follow-ups",
            section_style,
        )
    )

    followup_queryset = MedicalFollowUp.objects.select_related(
        "visit",
        "visit__player",
        "visit__team",
    ).filter(
        status=False
    )

    if filter_form.is_valid():

        if start_date:
            followup_queryset = followup_queryset.filter(
                visit__date__gte=start_date
            )

        if end_date:
            followup_queryset = followup_queryset.filter(
                visit__date__lte=end_date
            )

        if selected_team:
            followup_queryset = followup_queryset.filter(
                visit__team=selected_team
            )

        if selected_player:
            followup_queryset = followup_queryset.filter(
                visit__player=selected_player
            )

    followup_data = [
        [
            "Player",
            "Team",
            "Review Date",
            "Status",
        ]
    ]

    for followup in followup_queryset.order_by("review_date"):

        followup_data.append([
            display_value(followup.visit.player),
            display_value(followup.visit.team),
            display_value(followup.review_date),
            "Pending",
        ])

    if len(followup_data) == 1:
        followup_data.append([
            "No pending follow-ups",
            "-",
            "-",
            "-",
        ])

    story.append(
        make_table(
            followup_data,
            col_widths=[
                55 * mm,
                45 * mm,
                40 * mm,
                35 * mm,
            ],
        )
    )

    # ==========================================================
    # EXPECTED RETURNS
    # ==========================================================

    story.append(
        Paragraph(
            "Expected Player Returns",
            section_style,
        )
    )

    return_data = [
        [
            "Player",
            "Team",
            "Expected Return",
            "Availability",
        ]
    ]

    for visit in context["upcoming_returns"]:

        return_data.append([
            display_value(visit.player),
            display_value(visit.team),
            display_value(visit.expected_return_date),
            visit.get_availability_status_display(),
        ])

    if len(return_data) == 1:
        return_data.append([
            "No upcoming returns",
            "-",
            "-",
            "-",
        ])

    story.append(
        make_table(
            return_data,
            col_widths=[
                55 * mm,
                45 * mm,
                40 * mm,
                35 * mm,
            ],
        )
    )

    # ==========================================================
    # PAGE BREAK
    # ==========================================================

    story.append(PageBreak())

    # ==========================================================
    # DETAILED MEDICAL RECORDS
    # ==========================================================

    story.append(
        Paragraph(
            "Detailed Medical Records",
            section_style,
        )
    )

    medical_data = [
        [
            "Date",
            "Team",
            "Player",
            "Visit",
            "Complaint",
            "Status",
            "Training",
            "Availability",
        ]
    ]

    for visit in context["medical_visits"]:

        medical_data.append([
            display_value(visit.date),
            display_value(visit.team),
            display_value(visit.player),
            display_value(
                visit.get_visit_type_display()
            ),
            display_value(
                visit.get_main_complaint_display()
            ),
            display_value(
                visit.get_injury_status_display()
            ),
            display_value(
                visit.get_training_session_status_display()
            ),
            display_value(
                visit.get_availability_status_display()
            ),
        ])

    if len(medical_data) == 1:
        medical_data.append([
            "No medical records",
            "-",
            "-",
            "-",
            "-",
            "-",
            "-",
            "-",
        ])

    # Landscape table for detailed records
    detail_table = Table(
        medical_data,
        colWidths=[
            18 * mm,
            25 * mm,
            32 * mm,
            25 * mm,
            30 * mm,
            22 * mm,
            28 * mm,
            28 * mm,
        ],
        repeatRows=1,
    )

    detail_table.setStyle(
        TableStyle([
            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.HexColor("#1F4E78"),
            ),
            (
                "TEXTCOLOR",
                (0, 0),
                (-1, 0),
                colors.white,
            ),
            (
                "FONTNAME",
                (0, 0),
                (-1, 0),
                "Helvetica-Bold",
            ),
            (
                "FONTSIZE",
                (0, 0),
                (-1, -1),
                6,
            ),
            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.3,
                colors.grey,
            ),
            (
                "VALIGN",
                (0, 0),
                (-1, -1),
                "TOP",
            ),
            (
                "LEFTPADDING",
                (0, 0),
                (-1, -1),
                3,
            ),
            (
                "RIGHTPADDING",
                (0, 0),
                (-1, -1),
                3,
            ),
            (
                "TOPPADDING",
                (0, 0),
                (-1, -1),
                3,
            ),
            (
                "BOTTOMPADDING",
                (0, 0),
                (-1, -1),
                3,
            ),
        ])
    )

    story.append(detail_table)

    # ==========================================================
    # BUILD PDF
    # ==========================================================

    doc.build(story)

    pdf = buffer.getvalue()

    buffer.close()

    response = HttpResponse(
        pdf,
        content_type="application/pdf",
    )

    response[
        "Content-Disposition"
    ] = 'attachment; filename="medical_report.pdf"'

    return response