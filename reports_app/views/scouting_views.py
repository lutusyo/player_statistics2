import io
import pandas as pd
from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse
from django.utils import timezone
from openpyxl.styles import Alignment, Font
from openpyxl.drawing.image import Image as XLImage
from reports_app.models import Scouting
from reports_app.forms import ReportFilterForm
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from teams_app.models import Team


# ---- Display scouting reports ----
def scouting_reports(request, team_id):
    team = get_object_or_404(Team, id=team_id)
    form = ReportFilterForm(request.GET or None)
    queryset = Scouting.objects.filter(squad_id=team_id).select_related('squad')

    # Filtering
    if form.is_valid():
        team_filter = form.cleaned_data.get('team')
        period = form.cleaned_data.get('period')
        start_date = form.cleaned_data.get('start_date')
        end_date = form.cleaned_data.get('end_date')

        if team_filter:
            queryset = queryset.filter(squad=team_filter)
        if start_date and end_date:
            queryset = queryset.filter(date__range=[start_date, end_date])
        elif period and period != 'All':
            now = timezone.now().date()
            if period == 'This Week':
                queryset = queryset.filter(date__week=now.isocalendar()[1])
            elif period == 'This Month':
                queryset = queryset.filter(date__month=now.month)
            elif period == 'This Year':
                queryset = queryset.filter(date__year=now.year)

    return render(request, 'reports_app/2scouting/scouting_reports.html', {
        'form': form,
        'team': team,
        'records': queryset,
        'team_id': team_id,
    })


# ---- Export scouting reports to Excel ----
def export_scouting_excel(request, team_id):
    team = get_object_or_404(Team, id=team_id)
    queryset = Scouting.objects.filter(squad_id=team_id).select_related('squad')

    # Include all new fields
    df = pd.DataFrame(list(queryset.values(
        'name', 'pos', 'dob', 'squad__name', 'agreement', 'phone_number',
        'former_club', 'school_name', 'education_level', 'parent_or_coach_phone',
        'guardian_name', 'scouting_location', 'comments', 'date'
    )))

    # Create Excel
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Scouting Reports')
        ws = writer.sheets['Scouting Reports']

        # Add title
        ws.insert_rows(1)
        ws.insert_rows(2)
        ws.merge_cells('A1:M1')
        ws['A1'] = f"{team.name.upper()} - SCOUTING REPORTS"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

        # ✅ Fix: handle merged cells safely when formatting
        for col in ws.iter_cols(min_row=3):  # start from row 3 (header row)
            first_cell = col[0]
            col_letter = first_cell.column_letter
            max_length = 0
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
                cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.column_dimensions[col_letter].width = max_length + 3

        # Optionally add team logo (if exists)
        try:
            logo_path = f"static/images/{team.name.lower().replace(' ', '_')}_logo.png"
            img = XLImage(logo_path)
            img.height = 80
            img.width = 80
            ws.add_image(img, "A1")
        except Exception:
            pass

    buffer.seek(0)
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="scouting_reports.xlsx"'
    return response



# ---- Export scouting reports to PDF ----
def export_scouting_pdf(request, team_id):
    team = get_object_or_404(Team, id=team_id)
    queryset = Scouting.objects.filter(squad_id=team_id).select_related('squad')

    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    y = height - 50

    p.setFont("Helvetica-Bold", 14)
    p.drawString(180, y, f"{team.name.upper()} SCOUTING REPORT")
    y -= 40
    p.setFont("Helvetica", 10)

    for record in queryset:
        if y < 100:
            p.showPage()
            y = height - 50
            p.setFont("Helvetica", 10)
        p.drawString(50, y, f"{record.name} | {record.pos} | {record.squad.name} | {record.agreement} | {record.date}")
        y -= 20
        p.drawString(70, y, f"Phone: {record.phone_number or '-'} | Former Club: {record.former_club}")
        y -= 15
        p.drawString(70, y, f"School: {record.school_name or '-'} | Level: {record.education_level or '-'}")
        y -= 15
        p.drawString(70, y, f"Guardian: {record.guardian_name or '-'} | Contact: {record.parent_or_coach_phone or '-'}")
        y -= 25

    p.save()
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="scouting_reports.pdf"'
    return response
