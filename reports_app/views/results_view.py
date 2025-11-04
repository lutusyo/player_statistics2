from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse, Http404
from django.utils import timezone
from django.template.loader import render_to_string, get_template
from django.db.models import Q
from io import BytesIO
import datetime
import calendar
from matches_app.models import Match
from tagging_app.models import AttemptToGoal
from teams_app.models import Team

# For Excel
import openpyxl
from openpyxl.utils import get_column_letter

# ✅ For PDF (xhtml2pdf)
from xhtml2pdf import pisa


def parse_date(s):
    try:
        return datetime.datetime.strptime(s, "%Y-%m-%d").date()
    except Exception:
        return None


def get_date_range_from_request(request):
    """Handles period filters for matches."""
    period = request.GET.get('period', 'all')
    team_id = request.GET.get('team_id')  # optional

    today = timezone.localdate()

    if period == 'match':
        match_id = request.GET.get('match_id')
        if not match_id:
            raise Http404("match_id required for period=match")
        match = get_object_or_404(Match, pk=match_id)
        qs = Match.objects.filter(pk=match.pk)
        if team_id:
            qs = qs.filter(Q(home_team_id=team_id) | Q(away_team_id=team_id))
        return match.date, match.date, qs

    if period == 'week':
        start_str = request.GET.get('start')
        start = parse_date(start_str) or (today - datetime.timedelta(days=7))
        end = parse_date(request.GET.get('end')) or today
        qs = Match.objects.filter(date__range=(start, end))
    elif period == 'month':
        year = request.GET.get('year')
        month = request.GET.get('month')
        y = int(year) if year else today.year
        m = int(month) if month else today.month
        start = datetime.date(y, m, 1)
        last_day = calendar.monthrange(y, m)[1]
        end = datetime.date(y, m, last_day)
        qs = Match.objects.filter(date__range=(start, end))
    else:
        # all matches
        start = end = None
        qs = Match.objects.all()

    # ✅ Apply team filter if provided
    if team_id:
        qs = qs.filter(Q(home_team_id=team_id) | Q(away_team_id=team_id))

    return start, end, qs.order_by('date')



def compute_score_for_match(match):
    """Compute scoreline using AttemptToGoal model."""
    home = match.home_team
    away = match.away_team

    goals_home = AttemptToGoal.objects.filter(match=match, team=home, outcome='On Target Goal').count()
    goals_away = AttemptToGoal.objects.filter(match=match, team=away, outcome='On Target Goal').count()

    own_for_home = AttemptToGoal.objects.filter(match=match, is_own_goal=True, own_goal_for=home).count()
    own_for_away = AttemptToGoal.objects.filter(match=match, is_own_goal=True, own_goal_for=away).count()

    home_goals = goals_home + own_for_home
    away_goals = goals_away + own_for_away
    return home_goals, away_goals


def get_goal_scorers_for_match(match):
    """Return goal scorers and minute details."""
    qs = AttemptToGoal.objects.filter(match=match).filter(
        Q(outcome='On Target Goal') | Q(is_own_goal=True)
    ).select_related('player', 'team', 'own_goal_for').order_by('minute', 'second')

    scorers = []
    for a in qs:
        if a.is_own_goal:
            benef = a.own_goal_for.name if a.own_goal_for else 'Unknown'
            player_name = a.player.name if a.player else 'Unknown'
            note = f"OG (benefited {benef})"
            team_name = a.team.name if a.team else 'Unknown'
        else:
            team_name = a.team.name if a.team else 'Unknown'
            player_name = a.player.name if a.player else 'Unknown'
            note = ''
        scorers.append({'team': team_name, 'player': player_name, 'minute': a.minute, 'note': note})
    return scorers


def build_results_queryset_with_details(matches_qs, team_id=None):
    """Prepare rows for template and export."""
    rows = []
    for match in matches_qs:
        home_goals, away_goals = compute_score_for_match(match)
        result = f"{home_goals} - {away_goals}"
        scorers = get_goal_scorers_for_match(match)

        if team_id:
            try:
                team_id_int = int(team_id)
            except Exception:
                team_id_int = None
            if team_id_int == match.home_team_id:
                opponent = match.away_team.name
            elif team_id_int == match.away_team_id:
                opponent = match.home_team.name
            else:
                opponent = f"{match.home_team.name} vs {match.away_team.name}"
        else:
            opponent = f"{match.home_team.name} vs {match.away_team.name}"

        rows.append({
            'date': match.date,
            'opponent': opponent,
            'venue': match.venue,
            'competition_type': match.competition_type,
            'result': result,
            'scorers': scorers,
            'match_id': match.pk,
        })
    return rows


def results_list_view(request):
    """Main results list view with optional team filter."""
    start, end, matches_qs = get_date_range_from_request(request)
    team_id = request.GET.get('team_id')

    # ✅ Filter matches by selected team (either home_team or away_team)
    if team_id:
        matches_qs = matches_qs.filter(
            Q(home_team_id=team_id) | Q(away_team_id=team_id)
        )

    rows = build_results_queryset_with_details(matches_qs, team_id=team_id)
    teams = Team.objects.all().order_by('name')  # for dropdown filter

    context = {
        'rows': rows,
        'teams': teams,
        'period': request.GET.get('period', 'all'),
        'start': start,
        'end': end,
        'team_id': team_id,
        'query_params': request.GET.urlencode(),
    }
    return render(request, 'reports_app/results/results_list.html', context)




# ----------------- EXPORTS -----------------


def results_export_excel(request):
    """Export match results to Excel."""
    start, end, matches_qs = get_date_range_from_request(request)
    team_id = request.GET.get('team_id')
    rows = build_results_queryset_with_details(matches_qs, team_id=team_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"

    headers = ['Date', 'Opponent', 'Venue', 'Competition', 'Result', 'Goal Scorers']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    for r, row in enumerate(rows, start=2):
        scorers_text = "; ".join(
            f"{s['player']} ({s['minute']}'{(' - '+s['note']) if s['note'] else ''}) [{s['team']}]"
            for s in row['scorers']
        ) or '-'
        ws.cell(row=r, column=1, value=str(row['date']))
        ws.cell(row=r, column=2, value=row['opponent'])
        ws.cell(row=r, column=3, value=row['venue'])
        ws.cell(row=r, column=4, value=row['competition_type'])
        ws.cell(row=r, column=5, value=row['result'])
        ws.cell(row=r, column=6, value=scorers_text)

    for col in ws.columns:
        max_length = 0
        column = col[0].column
        for cell in col:
            val = str(cell.value) if cell.value else ""
            if len(val) > max_length:
                max_length = len(val)
        ws.column_dimensions[get_column_letter(column)].width = min(max_length + 2, 60)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"match_results_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    resp = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


def results_export_pdf(request):
    """Export match results to PDF using xhtml2pdf."""
    start, end, matches_qs = get_date_range_from_request(request)
    team_id = request.GET.get('team_id')
    rows = build_results_queryset_with_details(matches_qs, team_id=team_id)

    template = get_template('reports_app/results/results_pdf.html')
    html = template.render({
        'rows': rows,
        'period': request.GET.get('period', 'all'),
        'start': start,
        'end': end,
    })

    response = HttpResponse(content_type='application/pdf')
    filename = f"match_results_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('Error generating PDF', status=500)
    return response
