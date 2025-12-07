import io
import os
from io import BytesIO
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment
from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse
from django.utils import timezone
from django.template.loader import render_to_string
from weasyprint import HTML, CSS
from reports_app.models import Result
from reports_app.forms import ReportFilterForm
from teams_app.models import Team, StaffMember

from django.conf import settings
import os


# ---- Display result reports ----
def result_reports(request, team_id):
    team = get_object_or_404(Team, id=team_id)
    form = ReportFilterForm(request.GET or None)
    queryset = Result.objects.filter(our_team_id=team_id).select_related('home_team', 'away_team', 'our_team')

    if form.is_valid():
        team = form.cleaned_data.get('team')
        if team:
            queryset = queryset.filter(our_team=team)

    return render(request, 'reports_app/daily_report_templates/8results/result_reports.html', {
        'form': form,
        'team': team,
        'records': queryset,
        'team_id': team_id,
    })


# ---- Export to Excel ----
def export_results_excel(request, team_id):
    """Generate Excel file with team info and results."""
    team = get_object_or_404(Team, pk=team_id)
    results = Result.objects.filter(our_team=team).order_by('date')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"

    # --- Header with logo & team name ---
    ws.merge_cells('A1:H1')
    ws['A1'] = f"{team.name} Results Report"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws['A2'] = f"Age Group: {team.age_group or '-'}"
    ws['D2'] = f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M')}"

    # Add logo if it exists
    if team.logo and os.path.exists(team.logo.path):
        img = XLImage(team.logo.path)
        img.width = 80
        img.height = 80
        ws.add_image(img, 'H1')

    # Table headers
    headers = ['Date', 'Venue', 'Competition', 'Opponent', 'Score', 'Result', 'Goal Scorers', 'Notes']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Data rows
    for i, r in enumerate(results, start=5):
        opponent = r.away_team.name if r.our_team == r.home_team else r.home_team.name
        ws.cell(row=i, column=1, value=str(r.date))
        ws.cell(row=i, column=2, value=r.venue)
        ws.cell(row=i, column=3, value=r.competition_type)
        ws.cell(row=i, column=4, value=opponent)
        ws.cell(row=i, column=5, value=f"{r.home_score}-{r.away_score}")
        ws.cell(row=i, column=6, value=r.get_result_display())
        ws.cell(row=i, column=7, value=r.goal_scorers)
        ws.cell(row=i, column=8, value=r.notes)

    # Auto column width
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col[0].column)].width = min(max_len + 2, 40)

    # Save & return
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{team.name.replace(' ', '_')}_Results_{timezone.now().strftime('%Y%m%d')}.xlsx"
    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ---- Export to PDF ----
def export_results_pdf(request, team_id):
    """Generate a professional team results report PDF."""
    team = get_object_or_404(Team, pk=team_id)
    results = Result.objects.filter(our_team=team).order_by('date')
    staff_members = StaffMember.objects.filter(age_group=team.age_group)



    footer_image_path = os.path.join(settings.STATIC_ROOT, "reports_app/daily_report_statics/images/footer.png")

    context = {
        'team': team,
        'results': results,
        'staff_members': staff_members,
        'generated_on': timezone.now(),
        "footer_image_path": "file://" + footer_image_path, 
    }

    html_string = render_to_string('reports_app/daily_report_templates/8results/results_pdf.html', context)
    html = HTML(string=html_string, base_url=request.build_absolute_uri('/'))

    pdf_file = BytesIO()
    html.write_pdf(
        target=pdf_file,
        stylesheets=[CSS(string="""
            @page { size: A4; margin: 1cm; }
            body { font-family: 'Arial', sans-serif; font-size: 12px; color: #222; }
            h1, h2, h3 { text-align: center; color: #003366; margin-bottom: 6px; }
            table { width: 100%; border-collapse: collapse; margin-top: 10px; }
            th, td { border: 1px solid #ccc; padding: 5px; font-size: 11px; }
            th { background: #e8f0ff; font-weight: bold; }
            .team-header { text-align: center; margin-bottom: 15px; }
            .team-logo { width: 90px; height: 90px; margin-bottom: 10px; }
            .staff-table { margin-top: 15px; border: none; }
            .staff-table td { border: none; padding: 4px 8px; font-size: 11px; }
        """)]
    )

    pdf_file.seek(0)
    filename = f"{team.name.replace(' ', '_')}_Results_{timezone.now().strftime('%Y%m%d')}.pdf"
    response = HttpResponse(pdf_file.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
