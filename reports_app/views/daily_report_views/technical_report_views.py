import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, render
from django.utils.dateparse import parse_date

from teams_app.models import Team
from reports_app.models import Result, Medical, Transition, Scouting, Performance, IndividualActionPlan
from reports_app.views.daily_report_views.statistics_view import get_statistics_report
from players_app.models import Player
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def get_full_name(player):
    """Return the full player name as: First + Second + Surname"""
    if not player:
        return ""
    return f"{player.name} {player.second_name} {player.surname}".strip()



def write_sheet(ws, headers, rows, sum_columns=None):
    """
    Write headers and rows to worksheet with styling and auto-width.
    sum_columns: list of column indices (0-based) to sum at the bottom, e.g., [8,9,10] for Goals, Assists, Pre-Assists
    """
    sum_columns = sum_columns or []

    # Header styling
    header_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12, underline="single")  # bold, size 12, underline
    red_border = Border(bottom=Side(style="thin", color="FF0000"))  # red underline

    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = red_border

    # Add rows and center them
    for row in rows:
        ws.append(row)
    for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row_cells:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Add total row if sum_columns provided
    if sum_columns and rows:
        totals = ["TOTAL" if i == 0 else "" for i in range(len(headers))]
        for col_idx in sum_columns:
            totals[col_idx] = sum(row[col_idx] if isinstance(row[col_idx], (int, float)) else 0 for row in rows)
        ws.append(totals)
        # Style total row
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Auto-adjust column widths
    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            if cell.value:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
        ws.column_dimensions[column].width = min(max_length + 5, 50)  # max width 50




def download_technical_report(request, team_id):
    team = get_object_or_404(Team, id=team_id)
    season = request.GET.get("season", "")

    start_date_str = request.GET.get("start_date")
    end_date_str = request.GET.get("end_date")
    start_date = parse_date(start_date_str) if start_date_str else None
    end_date = parse_date(end_date_str) if end_date_str else None

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    def filter_by_date(queryset, date_field='date'):
        if start_date:
            queryset = queryset.filter(**{f"{date_field}__gte": start_date})
        if end_date:
            queryset = queryset.filter(**{f"{date_field}__lte": end_date})
        if season and hasattr(queryset.model, 'season'):
            queryset = queryset.filter(season=season)
        return queryset

    # ================ RESULTS ================
    ws = wb.create_sheet("Results")
    results = filter_by_date(Result.objects.filter(our_team=team))
    
    results_rows = []
    for r in results:
        # Format goal scorers nicely if they are stored as "PlayerName 23, PlayerName 45"
        goal_scorers_text = ""
        if r.goal_scorers:
            scorers_list = [s.strip() for s in r.goal_scorers.split(",")]
            full_names = []
            for s in scorers_list:
                player_name_part = s.rsplit(" ", 1)[0]  # remove minute
                player_obj = Player.objects.filter(name__iexact=player_name_part).first()
                if player_obj:
                    full_names.append(f"{get_full_name(player_obj)} {s.rsplit(' ', 1)[1]}")
                else:
                    full_names.append(s)
            goal_scorers_text = ", ".join(full_names)
        
        results_rows.append([
            r.date,
            r.competition_type,
            r.home_team.name,
            f"{r.home_score}-{r.away_score}",
            r.away_team.name,
            r.result,
            goal_scorers_text
        ])
    
    write_sheet(
        ws,
        ["Date", "Competition", "Home", "Score", "Away", "Result", "Goal Scorers"],
        results_rows
    )

    # ================ MEDICAL ================
    ws = wb.create_sheet("Medical")
    medicals = filter_by_date(Medical.objects.filter(squad=team))
    write_sheet(
        ws,
        ["Player", "Date", "Injury / Illness", "Status", "Comments"],
        [[get_full_name(m.name), m.date, m.injury_or_illness, m.status, m.comments] for m in medicals]
    )

    # ================ TRANSITION ================
    ws = wb.create_sheet("Transition")
    transitions = filter_by_date(Transition.objects.filter(squad=team))
    write_sheet(
        ws,
        ["Player", "Activity", "Played For", "Comments", "Date"],
        [[get_full_name(t.name), t.activity, t.played_for, t.comments, t.date] for t in transitions]
    )

    # ================ SCOUTING ================
    ws = wb.create_sheet("Scouting")
    scouting = filter_by_date(Scouting.objects.filter(squad=team))
    write_sheet(
        ws,
        ["Name", "DOB", "Position", "Agreement", "Former Club", "Comments"],
        [[s.name, s.dob, s.pos, s.agreement, s.former_club, s.comments] for s in scouting]
    )

    # ================ PERFORMANCE ================
    ws = wb.create_sheet("Performance")
    performances = filter_by_date(Performance.objects.filter(squad=team))
    write_sheet(
        ws,
        ["Date", "Activity", "Comments"],
        [[p.date, p.activity, p.comments] for p in performances]
    )

    # ================ INDIVIDUAL ACTION PLAN ================
    ws = wb.create_sheet("Individual Action Plan")
    iaps = filter_by_date(IndividualActionPlan.objects.filter(squad=team))
    write_sheet(
        ws,
        ["Player", "Category", "Responsibility", "Action", "Status", "Follow Up"],
        [[get_full_name(i.name), i.category, i.responsibility, i.action, i.status, i.follow_up] for i in iaps]
    )


    # ================ STATISTICS ================
    ws = wb.create_sheet("Statistics")
    stats_report = get_statistics_report(
        filter_type=request.GET.get("filter", "all"),
        team=team,
        start_date=start_date,
        end_date=end_date
    )

    # Add "PRE-ASSISTS" column
    headers = [
        "NAME", "POS", "TRAINING MINS", "GAME MINS", "APPS", "STARTS",
        "SUB IN", "SUB OUT", "GOALS", "ASSISTS", "PRE-ASSISTS", "NOTE"
    ]
    stats_rows = []
    for r in stats_report:
        stats_rows.append([
            get_full_name(r["player"]),
            r["position"],
            r["training_minutes"],
            r["game_minutes"],
            r["appearances"],
            r["starts"],
            r["sub_in"],
            r["sub_out"],
            r["goals"],
            r["assists"],
            r.get("pre_assists", 0),  # <-- Pre-Assists added here
            r["note"],
        ])

    sum_cols = [8, 9, 10]    

    write_sheet(ws, headers, stats_rows)


    # ================ DOWNLOAD ================
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"Technical_Report_{team.age_group.code}_{season or 'ALL'}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response
